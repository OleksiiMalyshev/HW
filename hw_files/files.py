import pickle
import openpyxl
import numpy

# TASK 1
# Opening file task1

with open("task1.txt", 'r') as file:
    lines = file.readlines()
    lines = [line.rstrip() for line in lines]
    subtitles = dict(zip(lines[::2], lines[1::2]))  # Dict creation

print(subtitles)
with open('new_file.txt', 'r+') as new_file:  # Opening new file
    for i in subtitles:
        new_file.write(subtitles[i])  # Adding to new file only text without time codes

# TASK 2
with open("task2", "rb") as file_2:
    print(numpy.mean(pickle.load(file_2)))  # Prints mean from task2 file list


# TASK 3
class excel_manager:
    def __init__(self, path_name):
        self.path_name = path_name
        self.excel_book = openpyxl.load_workbook(self.path_name)

    def __enter__(self):
        return self.excel_book

    def __exit__(self, exc_type, exc_val, exc_tb):
        if exc_type is None:
            self.excel_book.save(self.path_name)
            self.excel_book.close()
        else:
            print("\nSomething gone wrong! Please rewrite the data in Excel Manager")
            self.excel_book.close()
            return True


with excel_manager("file_excel.xlsx") as manager:
    active_sheet = manager.active
    print(manager.sheetnames)
    text = active_sheet.cell(1, 1)
    print(active_sheet['A1'].value)
    text.value = "Lol_it_works!"
    print(active_sheet.cell(2, 2, 'Another cell').value)
    raise Exception
